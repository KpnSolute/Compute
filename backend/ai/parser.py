"""File parser — converts uploaded files to structured rows, plain text, or invoice items."""

import ast
import csv
import io
import re
from typing import Any

from backend.ai import invoice_parser
from backend import inventory_formulas as fi


def _num(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int | float):
        return value
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _clamp_nonneg(value: Any) -> float | int | None:
    """Clamp a numeric value to >= 0, preserving None.

    Physical stock and weekly receive/issue counts can never be negative, but
    MJCC fact-check workbooks legitimately carry negative *ending* balances as an
    audit signal (more pulled than received+starting). The dispatch layer rejects
    negatives outright, which aborts the whole commit. We resolve this at data
    entry — the value is floored to 0 here so a single audit artifact can't block
    importing the other 191 items. The original figure remains visible in the
    source workbook's audit tabs.
    """
    n = _num(value)
    if n is None:
        return None
    return n if n >= 0 else 0


def _signed_num(value: Any) -> float | int | None:
    """Parse a signed numeric value.

    Quantity cells are non-negative physical counts, but Review financial flow
    cells can be negative when price/cost movement makes ending value exceed
    opening plus receipts. Preserve those signed audit values.
    """
    return _num(value)


def _inventory_category(label: Any) -> str | None:
    text = str(label or "").strip()
    if not text:
        return None
    norm = re.sub(r"[^a-z0-9]", "", text.lower())
    if norm in {"miamijobcorpscafeteriainventory", "itemdescription"}:
        return None
    if "total" in norm:
        return None
    # Maps normalized cell text from MJCC monthly inventory workbooks to exact
    # inventory_categories names. Taxonomy: Dairy, Cereal, Beverages, Snacks,
    # Meats, Frozen Food, Dry Goods, Produce, Disposables.
    category_map = {
        "dairy": "Dairy",
        "cereal": "Cereal",
        "beverages": "Beverages",
        "beverage": "Beverages",
        "snacks": "Snacks",
        "snack": "Snacks",
        "meats": "Meats",
        "meat": "Meats",
        "protein": "Meats",
        "frozenfood": "Frozen Food",
        "frozenfoods": "Frozen Food",
        "frozengoods": "Frozen Food",
        "frozen": "Frozen Food",
        "drygoods": "Dry Goods",
        "dry": "Dry Goods",
        "produce": "Produce",
        "fresh": "Produce",
        "disposibles": "Disposables",
        "disposables": "Disposables",
        "supplies": "Disposables",
        "supply": "Disposables",
    }
    exact = category_map.get(norm)
    if exact:
        return exact
    # Substring fallback for compound labels like "Produce & Fresh" or
    # "Protein & Meat" (normalize → "producefresh" / "proteinmeat") that don't
    # match an exact key. Order matters — most specific first.
    for token, canonical in (
        ("frozen", "Frozen Food"),
        ("produce", "Produce"),
        ("protein", "Meats"),
        ("meat", "Meats"),
        ("dairy", "Dairy"),
        ("cereal", "Cereal"),
        ("beverage", "Beverages"),
        ("snack", "Snacks"),
        ("drygood", "Dry Goods"),
        ("disposable", "Disposables"),
        ("suppl", "Disposables"),
    ):
        if token in norm:
            return canonical
    return None


def _inventory_sku(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0)?", text):
        digits = text.split(".", 1)[0]
        if len(digits) < 4:
            return ""
        return digits
    return text


def _find_mjcc_grid_header(
    rows: list[tuple],
) -> tuple[int, int] | None:
    """Locate the MJCC inventory header row and description column.

    Scans every row looking for "item description" (case-insensitive) in any
    of the first 5 columns.  Returns (row_index, desc_col_index) or None.

    The original parser required "item description" at exactly column B (index 1)
    with issued/received headers at fixed offsets.  That fails for workbooks where
    the layout is shifted (column A used for SKU, item-description in column B is
    fine, but some "fact-checked" variants put the description in column A or C).
    """
    for row_idx, row in enumerate(rows):
        for col_idx in range(min(5, len(row))):
            cell = str(row[col_idx] or "").strip().lower()
            if cell == "item description" or cell == "description":
                return row_idx, col_idx
    return None


def _parse_mjcc_monthly_inventory(content: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parsed: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))

        header_info = _find_mjcc_grid_header(rows)
        if not header_info:
            continue
        header_row_idx, desc_col = header_info
        # New standard 14-col workbook has named headers like "Opening OH" /
        # "Received Wk1" — offset arithmetic below is wrong for that layout.
        # Bail out and let _parse_mjcc_flat_inventory handle it instead.
        _hdr = rows[header_row_idx]
        _hdr_norm = {re.sub(r"[^a-z0-9]", "", str(c or "").lower()) for c in _hdr}
        if "openingoh" in _hdr_norm or "receivedwk1" in _hdr_norm:
            continue

        # Derive column offsets relative to the description column.
        # Standard MJCC layout (desc_col=1):
        #   col 0 = SKU, col 1 = Item Description, col 2 = On Hand,
        #   col 3 = Unit Price, col 4 = Unit (sometimes absent),
        #   cols 5-8 = W1-W4 Issued, cols 9-12 = W1-W4 Received
        # When desc_col differs, shift all offsets accordingly.
        sku_col = desc_col - 1 if desc_col > 0 else None
        onhand_col = desc_col + 1
        price_col = desc_col + 2
        # Issued / received: offset from desc_col.  In the standard layout
        # there is one "gap" column (unit) at desc_col+3, then 4 issued columns,
        # then 4 received columns.  Use a flexible scan: treat any column at
        # desc_col+3 or later that is numeric as week data in order.
        w_start = desc_col + 3  # first weekly column (W1 issued or W1 received)
        cat_col = desc_col  # category labels appear in the same column as desc

        # Need at least onhand column to bother parsing
        min_cols_needed = price_col + 1
        if not any(len(row) >= min_cols_needed for row in rows):
            continue

        category: str | None = None
        for row in rows:
            cells = list(row) + [None] * max(0, w_start + 9 - len(row))
            maybe_category = _inventory_category(cells[cat_col])
            # Detect amount columns: onhand, price, and first four weekly cols
            amount_indices = [onhand_col, price_col] + list(range(w_start, w_start + 4))
            row_has_item_amounts = any(
                _num(cells[idx]) is not None
                for idx in amount_indices
                if idx < len(cells)
            )
            if maybe_category and not row_has_item_amounts:
                category = maybe_category
                continue
            if category is None:
                continue

            desc = str(cells[cat_col] or "").strip()
            if (
                not desc
                or desc.lower() in ("item description", "description")
                or "total" in desc.lower()
            ):
                continue
            onhand_val = (
                _clamp_nonneg(cells[onhand_col]) if onhand_col < len(cells) else None
            )
            price_val = (
                _clamp_nonneg(cells[price_col]) if price_col < len(cells) else None
            )
            sku_val = (
                _inventory_sku(cells[sku_col])
                if sku_col is not None and sku_col < len(cells)
                else ""
            )
            if onhand_val is None and price_val is None and not sku_val:
                continue

            def _wcol(offset: int) -> float | int | None:
                idx = w_start + offset
                return _clamp_nonneg(cells[idx]) if idx < len(cells) else None

            # Legacy grid: cols 5-8 = W1-W4 pulled, cols 9-12 = W1-W4 received.
            # New schema keeps 3 weeks; W4 is folded into W3 so historical W4
            # activity is not silently dropped on import.
            parsed.append(
                {
                    "sku": sku_val,
                    "desc": desc,
                    "category": category,
                    "onHand": onhand_val or 0,
                    "price": price_val,
                    "w1p": _wcol(0) or 0,
                    "w2p": _wcol(1) or 0,
                    "w3p": (_wcol(2) or 0) + (_wcol(3) or 0),
                    "w1r": _wcol(4) or 0,
                    "w2r": _wcol(5) or 0,
                    "w3r": (_wcol(6) or 0) + (_wcol(7) or 0),
                    "unit": "each",
                    "__sheet": ws.title,
                }
            )
    return parsed


_FLAT_INV_HEADER_ALIASES: dict[str, str] = {
    # category
    "category": "category",
    "cat": "category",
    # sku
    "sku": "sku",
    "originalsku": "sku",
    "invoicesku": "sku",
    "itemcode": "sku",
    "code": "sku",
    # description
    "description": "desc",
    "itemdescription": "desc",
    "invoicedescription": "desc",
    "item": "desc",
    # on-hand = the OPENING balance for the period. In this system on_hand is the
    # opening figure and ENDING is computed (opening + received - issued). So map
    # the STARTING balance here. Mapping "Ending OH" was a bug: it fed the closing
    # stock back in as the opening, which hid all received/issued activity and
    # double-stated the balance. "Ending OH" / "Ending Value" are deliberately
    # NOT mapped — ending is derived, never imported as on_hand.
    "startoh": "onHand",
    "startingoh": "onHand",
    "startonhand": "onHand",
    "startingonhand": "onHand",
    "startingbalance": "onHand",
    "beginningoh": "onHand",
    "openingoh": "onHand",
    "openingbalance": "onHand",
    # Plain single-column snapshots (no separate start/ending) — treat the lone
    # on-hand column as the opening for entry.
    "onhand": "onHand",
    "qtyonhand": "onHand",
    # price
    "unitprice": "price",
    "latestunitcost": "price",
    "latestunitcost$": "price",
    "unitcost": "price",
    "price": "price",
    # financial controls from the workbook Review sheet
    "openingunitcost": "opening_unit_cost",
    "openingvalue": "opening_value",
    "receivedvalue": "received_value",
    "pulledvalue": "pulled_value",
    "issuedvalue": "pulled_value",
    "inventoryflowvalue": "pulled_value",
    "endingvalue": "ending_value",
    # weekly received/pulled — Monthly Inventory Template (3 weeks)
    "receivedwk1": "w1r",
    "receivedwk2": "w2r",
    "receivedwk3": "w3r",
    "pulledwk1": "w1p",
    "pulledwk2": "w2p",
    "pulledwk3": "w3p",
    # May case: weekly pull cols blank but verified monthly Total Pulled present.
    # Mapped to a passthrough field so dispatch can apply it without inventing
    # per-week distribution. Total Received and Ending OH are NOT mapped.
    "totalpulled": "total_pulled_raw",
    # par / reorder threshold (header normalized: non-alphanumerics stripped)
    "par": "par",
    "parlevel": "par",
    "parstock": "par",
    "reorder": "par",
    "reorderpoint": "par",
    "reorderlevel": "par",
    "minstock": "par",
    "minimum": "par",
    "minqty": "par",
    "threshold": "par",
}


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _safe_arithmetic(expr: str) -> float | None:
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError:
        return None

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, int | float):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = _eval(node.operand)
            return value if isinstance(node.op, ast.UAdd) else -value
        if isinstance(node, ast.BinOp) and isinstance(
            node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)
        ):
            left = _eval(node.left)
            right = _eval(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if right == 0:
                return None
            return left / right
        raise ValueError("unsupported formula")

    try:
        return _eval(tree)
    except Exception:
        return None


def _formula_value(
    wb,
    formula_wb,
    formula: Any,
    current_sheet: str | None = None,
    depth: int = 0,
    cache: dict[tuple[str, int, int], Any] | None = None,
) -> Any:
    if not isinstance(formula, str) or not formula.startswith("=") or depth > 4:
        return None
    try:
        from openpyxl.utils.cell import column_index_from_string
    except ImportError:
        return None

    cache = cache if cache is not None else {}

    def _resolve_ref(token: str) -> Any:
        if "!" in token:
            sheet_part, cell_ref = token.split("!", 1)
            sheet = sheet_part.strip("'")
        else:
            sheet = current_sheet
            cell_ref = token
        if not sheet or sheet not in wb.sheetnames:
            return 0
        cell_match = re.fullmatch(r"\$?(?P<col>[A-Z]+)\$?(?P<row>\d+)", cell_ref, re.I)
        if not cell_match:
            return 0
        row = int(cell_match.group("row"))
        col = column_index_from_string(cell_match.group("col").upper())
        cache_key = (sheet, row, col)
        if cache_key in cache:
            return cache[cache_key]
        value = wb[sheet].cell(row=row, column=col).value
        formula = (
            formula_wb[sheet].cell(row=row, column=col).value
            if formula_wb is not None and sheet in formula_wb.sheetnames
            else None
        )
        if isinstance(formula, str) and formula.startswith("="):
            value = _formula_value(
                wb,
                formula_wb,
                formula,
                sheet,
                depth + 1,
                cache,
            )
        value = value if value is not None else 0
        cache[cache_key] = value
        return value

    expression = formula.strip()[1:].replace("$", "")
    direct_ref = re.fullmatch(
        r"(?:'[^']+'|[A-Z][A-Z0-9 _]*)![A-Z]+\d+|[A-Z]+\d+",
        expression,
        flags=re.IGNORECASE,
    )
    if direct_ref:
        return _resolve_ref(expression)

    def _range_values(token: str) -> list[Any]:
        if ":" not in token:
            return [_resolve_ref(token)]
        start_ref, end_ref = token.split(":", 1)
        sheet_prefix = ""
        if "!" in start_ref and "!" not in end_ref:
            sheet_prefix, start_ref = start_ref.split("!", 1)
            sheet_prefix += "!"
        start_match = re.fullmatch(r"([A-Z]+)(\d+)", start_ref, re.I)
        end_match = re.fullmatch(r"([A-Z]+)(\d+)", end_ref, re.I)
        if not start_match or not end_match:
            return []
        if start_match.group(1).upper() != end_match.group(1).upper():
            return []
        col = start_match.group(1).upper()
        sheet_name = sheet_prefix[:-1] if sheet_prefix else current_sheet
        if sheet_name and sheet_name.startswith("'") and sheet_name.endswith("'"):
            sheet_name = sheet_name[1:-1]
        start_row = int(start_match.group(2))
        end_row = int(end_match.group(2))
        values = []
        for row_idx in range(min(start_row, end_row), max(start_row, end_row) + 1):
            token_ref = f"{sheet_prefix}{col}{row_idx}"
            if (
                sheet_name
                and formula_wb is not None
                and sheet_name in formula_wb.sheetnames
            ):
                col_idx = column_index_from_string(col)
                formula = formula_wb[sheet_name].cell(row=row_idx, column=col_idx).value
                if isinstance(formula, str) and formula.startswith("="):
                    values.append(
                        _formula_value(wb, formula_wb, formula, sheet_name, cache={})
                    )
                    continue
            values.append(_resolve_ref(token_ref))
        return values

    countif_match = re.fullmatch(
        r'COUNTIF\((?P<range>[^,]+),(?P<criteria>"[^"]+"|[^)]+)\)',
        expression,
        flags=re.IGNORECASE,
    )
    if countif_match:
        criteria = countif_match.group("criteria").strip().strip('"')
        values = _range_values(countif_match.group("range").strip())
        if criteria == "TEMP*":
            return sum(1 for value in values if str(value or "").startswith("TEMP"))
        if criteria == "<>TEMP*":
            return sum(1 for value in values if not str(value or "").startswith("TEMP"))
        if criteria == "<0":
            return sum(1 for value in values if fi.num(value) < 0)
        return 0

    sum_match = re.fullmatch(r"SUM\((?P<args>[^)]+)\)", expression, flags=re.IGNORECASE)
    if sum_match:
        total = 0.0
        for token in sum_match.group("args").split(","):
            token = token.strip()
            if ":" in token:
                total += sum(fi.num(value) for value in _range_values(token))
                continue
            value = (
                _resolve_ref(token) if re.search(r"[A-Z]", token, re.I) else _num(token)
            )
            total += fi.num(value)
        return total
    if not re.fullmatch(r"[A-Z0-9_ '!+\-*/().]+", expression, flags=re.IGNORECASE):
        return None

    ref_re = re.compile(
        r"(?:'[^']+'|[A-Z][A-Z0-9 _]*)![A-Z]+\d+|[A-Z]+\d+",
        flags=re.IGNORECASE,
    )
    numeric_expr = ref_re.sub(
        lambda m: str(fi.num(_resolve_ref(m.group(0)))), expression
    )
    return _safe_arithmetic(numeric_expr)


def _parse_review_value_rows(wb, formula_wb=None) -> dict[str, dict[str, Any]]:
    """Extract audited financial values from the workbook Review sheet.

    The flat Inventory sheet has quantities and a current unit price. The Review
    sheet carries the auditable dollar values used by MJCC: opening value from
    the prior month, receipt value from invoices, pull/flow value, and ending
    value. Key these by SKU so the Inventory rows can be enriched without
    double-counting the Review sheet as a second import table.
    """
    value_by_sku: dict[str, dict[str, Any]] = {}
    for ws in wb.worksheets:
        if "review" not in str(ws.title or "").lower():
            continue
        formula_cache: dict[tuple[str, int, int], Any] = {}
        rows = list(ws.iter_rows(values_only=True))
        formula_rows = []
        if formula_wb is not None and ws.title in formula_wb.sheetnames:
            formula_rows = list(formula_wb[ws.title].iter_rows(values_only=True))
        header_idx: int | None = None
        col_map: dict[int, str] = {}
        for r_idx, row in enumerate(rows):
            mapping: dict[int, str] = {}
            for c_idx, cell in enumerate(row):
                canonical = _FLAT_INV_HEADER_ALIASES.get(_norm_header(cell))
                if canonical and c_idx not in mapping:
                    mapping[c_idx] = canonical
            fields = set(mapping.values())
            if {
                "sku",
                "desc",
                "opening_value",
                "received_value",
                "ending_value",
            }.issubset(fields):
                header_idx = r_idx
                col_map = mapping
                break
        if header_idx is None:
            continue
        for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 1):
            if all(v is None for v in row):
                continue
            formula_row = formula_rows[offset] if offset < len(formula_rows) else ()
            rec: dict[str, Any] = {}
            for c_idx, canonical in col_map.items():
                if c_idx < len(row):
                    value = row[c_idx]
                    if (
                        value is None
                        and formula_wb is not None
                        and c_idx < len(formula_row)
                    ):
                        value = _formula_value(
                            wb,
                            formula_wb,
                            formula_row[c_idx],
                            ws.title,
                            cache=formula_cache,
                        )
                    rec[canonical] = value
            sku = _inventory_sku(rec.get("sku"))
            desc = str(rec.get("desc") or "").strip()
            if not sku or not desc or "total" in desc.lower():
                continue
            values: dict[str, Any] = {}
            for key in (
                "opening_unit_cost",
                "opening_value",
                "received_value",
                "pulled_value",
                "ending_value",
            ):
                parsed = (
                    _signed_num(rec.get(key))
                    if key == "pulled_value"
                    else _clamp_nonneg(rec.get(key))
                )
                if parsed is not None:
                    values[key] = parsed
            if "pulled_value" not in values and (
                "opening_value" in values or "received_value" in values
            ):
                values["pulled_value"] = 0
            if "ending_value" not in values and (
                "opening_value" in values or "received_value" in values
            ):
                values["ending_value"] = fi.ending_value(
                    values.get("opening_value"),
                    values.get("received_value"),
                    values.get("pulled_value"),
                )
            if values:
                value_by_sku[sku] = values
    return value_by_sku


def _parse_mjcc_flat_inventory(content: bytes) -> list[dict[str, Any]]:
    """Parse a flat columnar MJCC inventory workbook (e.g. "* Full Inventory" /
    "* Fact checked" exports).

    Unlike the weekly grid (_parse_mjcc_monthly_inventory), these sheets carry a
    title banner in the top rows, then a single header row such as:

        Category | SKU | Description | Start OH | Total Rcvd | Total Pulled |
        Ending OH | Unit Price | Ending Value

    We locate that header row by scanning the first rows of each sheet for one
    that contains a description column plus a sku or category column, map columns
    by name, and emit one row per item.  Only the FIRST sheet with a usable
    header is consumed so multi-tab audit workbooks don't double-count.

    Mapping rules (see _FLAT_INV_HEADER_ALIASES):
    - Start OH  -> on_hand   (the OPENING balance; ending is computed downstream)
    - Unit Price -> unit_price
    - Ending OH / Ending Value are NOT imported — ending = opening + received -
      issued, so importing the closing figure as on_hand would double-state stock
      and hide all weekly activity.
    - Total Rcvd / Total Pulled are NOT forced into weekly columns — a monthly
      total has no honest week attribution, and the weekly received/issued data
      is owned by the per-week invoice/pull-sheet uploads. dispatch_inventory_save
      preserves existing weekly columns when those keys are omitted.
    """
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    formula_wb = openpyxl.load_workbook(
        io.BytesIO(content), read_only=False, data_only=False
    )
    review_values = _parse_review_value_rows(wb, formula_wb)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx: int | None = None
        col_map: dict[int, str] = {}
        # Scan the first 15 rows for a recognizable header.
        for r_idx, row in enumerate(rows[:15]):
            mapping: dict[int, str] = {}
            for c_idx, cell in enumerate(row):
                canonical = _FLAT_INV_HEADER_ALIASES.get(_norm_header(cell))
                if canonical and c_idx not in mapping:
                    mapping[c_idx] = canonical
            fields = set(mapping.values())
            if "desc" in fields and ("sku" in fields or "category" in fields):
                header_idx = r_idx
                col_map = mapping
                break

        if header_idx is None:
            continue

        parsed: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if all(v is None for v in row):
                continue
            rec: dict[str, Any] = {}
            for c_idx, canonical in col_map.items():
                if c_idx < len(row):
                    rec[canonical] = row[c_idx]

            desc = str(rec.get("desc") or "").strip()
            cat_raw = str(rec.get("category") or "").strip()
            # Skip total / summary / repeated-header rows.
            if not desc or "total" in desc.lower() or desc.lower() == "description":
                continue
            if "total" in cat_raw.lower():
                continue

            sku = _inventory_sku(rec.get("sku"))
            category = _inventory_category(cat_raw) or cat_raw or "Dry Goods"
            onhand = _clamp_nonneg(rec.get("onHand"))
            price = _clamp_nonneg(rec.get("price"))
            par = _clamp_nonneg(
                rec.get("par")
            )  # par/reorder level when the sheet has it
            if not sku and onhand is None and price is None:
                continue

            row_out: dict[str, Any] = {
                "sku": sku,
                "desc": desc,
                "category": category,
                "onHand": onhand if onhand is not None else 0,
                "price": price,
                "unit": "each",
                "__sheet": ws.title,
            }
            # Emit par for every row when the sheet HAS a par column (0 for blanks),
            # so map_rows_to_inventory sees the column even if the first row is blank.
            if "par" in col_map.values():
                row_out["par"] = par if par is not None else 0
            # Emit weekly received/pulled when the sheet carries those columns.
            for _wk in ("w1r", "w2r", "w3r", "w1p", "w2p", "w3p"):
                if _wk in col_map.values():
                    _v = _clamp_nonneg(rec.get(_wk))
                    row_out[_wk] = _v if _v is not None else 0
            # May case: weekly pull cols blank but Total Pulled is a verified monthly
            # figure. Preserve as total_pulled_raw so dispatch can apply it safely
            # without guessing per-week distribution.
            if "total_pulled_raw" in col_map.values():
                _weekly_pulls = sum(
                    row_out.get(k, 0) or 0 for k in ("w1p", "w2p", "w3p")
                )
                _tp = _clamp_nonneg(rec.get("total_pulled_raw"))
                if _weekly_pulls == 0 and _tp:
                    row_out["total_pulled_raw"] = _tp
            if sku in review_values:
                row_out.update(review_values[sku])
            parsed.append(row_out)

        if parsed:
            return parsed  # first usable sheet wins — avoid double-counting tabs
    return []


# Review-tab "Quantity Control" labels → reconciliation metric. The control block
# lists a verified total per metric (label in col A, value in the next column).
_REVIEW_CONTROL_LABELS = {
    "inventoryitems": "item_count",
    "invoiceskus": "invoice_skus",
    "openingtempitems": "temp_items",
    "openingoh": "opening",
    "totalreceived": "received",
    "totalpulled": "pulled",
    "endingoh": "ending",
    "negativeendingrows": "negative_ending_rows",
}


_REVIEW_FINANCIAL_LABELS = {
    "openinginventoryvalue": "opening_value",
    "productreceiptvalue": "received_value",
    "netinvoicereceiptvalue": "net_receipt_value",
    "creditssurchargenet": "credits_surcharge_net",
    "endinginventoryvalue": "ending_value",
    "inventoryflowvalue": "pulled_value",
    "endingdifferencecheck": "ending_difference_check",
}


def _grid_totals(content: bytes) -> dict[str, float] | None:
    """Sum the Inventory sheet into {item_count, opening, received, pulled, ending}.

    pulled uses the per-week Pulled columns, falling back to Total Pulled when the
    weekly cells are blank (May-style verified monthly total). Returns None if no
    recognizable Inventory grid is present.
    """
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return None
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = None
        cols: dict[str, int] = {}
        for i, row in enumerate(rows[:15]):
            norm = {_norm_header(c): j for j, c in enumerate(row)}
            if "openingoh" in norm and "receivedwk1" in norm:
                hdr_idx = i
                cols = norm
                break
        if hdr_idx is None:
            continue

        def _cell(row, key):
            j = cols.get(key)
            return _num(row[j]) if j is not None and j < len(row) else None

        items: list[dict[str, Any]] = []
        for row in rows[hdr_idx + 1 :]:
            if all(v is None for v in row):
                continue
            sku = row[cols["sku"]] if "sku" in cols and cols["sku"] < len(row) else None
            desc = (
                row[cols["description"]]
                if "description" in cols and cols["description"] < len(row)
                else None
            )
            d = str(desc or "").strip()
            if (
                (not sku and not d)
                or "total" in d.lower()
                or d.lower() in ("description", "item description")
            ):
                continue
            opening = fi.num(_cell(row, "openingoh"))
            received = fi.total_received(
                _cell(row, "receivedwk1"),
                _cell(row, "receivedwk2"),
                _cell(row, "receivedwk3"),
            )
            wk_pull = fi.total_pulled(
                _cell(row, "pulledwk1"),
                _cell(row, "pulledwk2"),
                _cell(row, "pulledwk3"),
            )
            # weekly pulls win; fall back to the verified monthly Total Pulled cell
            pulled = wk_pull if wk_pull else fi.num(_cell(row, "totalpulled"))
            items.append(
                {"sku": sku, "opening": opening, "received": received, "pulled": pulled}
            )
        return fi.review_controls(items)
    return None


def _review_controls(content: bytes) -> dict[str, float] | None:
    """Read the Review tab's Quantity Control block (verified totals per metric).

    Returns {item_count, opening, received, pulled, ending} from the verified
    column, or None when no Review control block is present.
    """
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=False, data_only=True
        )
        formula_wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=False, data_only=False
        )
    except Exception:
        return None
    for ws in wb.worksheets:
        if "review" not in str(ws.title or "").lower():
            continue
        formula_ws = formula_wb[ws.title] if ws.title in formula_wb.sheetnames else None
        controls: dict[str, float] = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            label = _norm_header(row[0])
            metric = _REVIEW_CONTROL_LABELS.get(label)
            if not metric or metric in controls:
                continue
            # Quantity Control verified total is column B.
            cell = row[1] if len(row) > 1 else None
            formula = (
                formula_ws.cell(row=row_idx, column=2).value
                if formula_ws is not None
                else None
            )
            val = None
            if isinstance(formula, str) and formula.startswith("="):
                val = _num(_formula_value(wb, formula_wb, formula, ws.title))
            if val is None:
                val = _num(cell)
            if val is not None:
                controls[metric] = val
        if controls:
            return controls
    return None


def _review_financial_controls(content: bytes) -> dict[str, float] | None:
    """Read the Review tab's Financial Control block."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=False, data_only=True
        )
        formula_wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=False, data_only=False
        )
    except Exception:
        return None
    for ws in wb.worksheets:
        if "review" not in str(ws.title or "").lower():
            continue
        formula_ws = formula_wb[ws.title] if ws.title in formula_wb.sheetnames else None
        controls: dict[str, float] = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for idx, cell in enumerate(row):
                metric = _REVIEW_FINANCIAL_LABELS.get(_norm_header(cell))
                if not metric or metric in controls:
                    continue
                value_idx = idx + 1
                value = row[value_idx] if value_idx < len(row) else None
                formula = (
                    formula_ws.cell(row=row_idx, column=value_idx + 1).value
                    if formula_ws is not None
                    else None
                )
                val = None
                if isinstance(formula, str) and formula.startswith("="):
                    val = _num(_formula_value(wb, formula_wb, formula, ws.title))
                if val is None:
                    val = _num(value)
                if val is not None:
                    controls[metric] = val
        if controls:
            return controls
    return None


def _review_weekly_invoice_totals(content: bytes) -> dict[str, Any] | None:
    """Read manager-entered weekly invoice totals from the Review tab.

    These are authoritative invoice product values. They are intentionally not
    derived from inventory quantities because invoices include real vendor
    product totals that can differ from quantity * catalog/unit price math.
    """
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return None

    week_label_re = re.compile(r"^week([1-5])invoicetotal$")
    for ws in wb.worksheets:
        if "review" not in str(ws.title or "").lower():
            continue
        weeks: dict[str, float] = {}
        notes: dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            match = week_label_re.match(_norm_header(row[0]))
            if not match:
                continue
            week = match.group(1)
            value = None
            for cell in row[1:6]:
                value = _num(cell)
                if value is not None:
                    break
            if value is None:
                continue
            weeks[week] = round(float(value), 2)
            for cell in row[2:]:
                note = str(cell or "").strip()
                if note and _num(note) is None:
                    notes[week] = note
                    break
        if weeks:
            total = round(sum(weeks.values()), 2)
            return {
                "source": "review_weekly_invoice_totals",
                "weeks": weeks,
                "total": total,
                "notes": notes,
            }
    return None


def _parsed_financial_totals(content: bytes) -> dict[str, float] | None:
    rows = _parse_mjcc_flat_inventory(content)
    if not rows:
        return None
    return {
        "opening_value": sum(fi.num(r.get("opening_value")) for r in rows),
        "received_value": sum(fi.num(r.get("received_value")) for r in rows),
        "pulled_value": sum(fi.num(r.get("pulled_value")) for r in rows),
        "ending_value": sum(fi.num(r.get("ending_value")) for r in rows),
    }


def extract_workbook_formula_report(content: bytes) -> dict[str, Any] | None:
    """Extract the workbook's derived-column formulas and recompute them internally.

    The template's Total Received / Total Pulled / Ending OH columns are formula
    driven (=SUM(E,G,I), =SUM(F,H,J), =D+K-L). Excel's cached results can be stale
    (saved without a recalc) or absent (workbook written programmatically), so the
    system must NOT trust those cached cells — it recomputes each derived column
    from the raw weekly cells (i.e. it applies the formula itself).

    Returns the extracted formula strings, whether they match the expected template
    formulas, and how many cached cells were stale (cached != recomputed). Returns
    None when there is no recognizable Inventory grid.
    """
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb_f = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
        wb_v = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None

    def _find(ws):
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:15]):
            norm = {_norm_header(c): j for j, c in enumerate(row)}
            if "openingoh" in norm and "receivedwk1" in norm:
                return i, norm
        return None, None

    ws_f = next((w for w in wb_f.worksheets if _find(w)[0] is not None), None)
    if ws_f is None:
        return None
    hdr_idx, cols = _find(ws_f)
    ws_v = wb_v[ws_f.title]
    first = hdr_idx + 2  # 1-based row of the first data row

    def _fstr(metric):
        j = cols.get(metric)
        if j is None:
            return None
        return ws_f.cell(row=first, column=j + 1).value

    formulas = {
        "total_received": _fstr("totalreceived"),
        "total_pulled": _fstr("totalpulled"),
        "ending_oh": _fstr("endingoh"),
    }
    # Expected template shapes (column letters are fixed in the template).
    expected = {
        "total_received": "=SUM(E,G,I)",
        "total_pulled": "=SUM(F,H,J)",
        "ending_oh": "=D+K-L",
    }

    def _shape(f):
        # strip row numbers so '=SUM(E2,G2,I2)' -> '=SUM(E,G,I)'
        return re.sub(r"\d+", "", str(f or "")).replace(" ", "").upper()

    template_match = {
        k: (_shape(formulas[k]) == expected[k]) for k in expected if formulas[k]
    }

    stale = {"total_received": 0, "total_pulled": 0, "ending_oh": 0}
    rows_v = list(ws_v.iter_rows(values_only=True))
    for vr in rows_v[hdr_idx + 1 :]:
        if all(c is None for c in vr):
            continue
        d = _num(vr[cols["openingoh"]]) or 0 if cols["openingoh"] < len(vr) else 0
        rec = sum(
            _num(vr[cols[k]]) or 0
            for k in ("receivedwk1", "receivedwk2", "receivedwk3")
            if cols.get(k) is not None and cols[k] < len(vr)
        )
        pul = sum(
            _num(vr[cols[k]]) or 0
            for k in ("pulledwk1", "pulledwk2", "pulledwk3")
            if cols.get(k) is not None and cols[k] < len(vr)
        )
        for metric, computed in (
            ("totalreceived", rec),
            ("totalpulled", pul),
            ("endingoh", d + rec - pul),
        ):
            j = cols.get(metric)
            if j is None or j >= len(vr):
                continue
            cached = _num(vr[j])
            if cached is not None and abs(cached - computed) > 0.001:
                key = {
                    "totalreceived": "total_received",
                    "totalpulled": "total_pulled",
                    "endingoh": "ending_oh",
                }[metric]
                stale[key] += 1

    return {
        "formulas": formulas,
        "template_match": template_match,
        "stale_cached_cells": stale,
        "recomputed_internally": True,
    }


def extract_workbook_reconciliation(content: bytes) -> dict[str, Any] | None:
    """Reconcile a monthly workbook's Inventory grid against its Review controls.

    The Inventory grid is the line-item detail; the Review tab's Quantity Control
    block is the manager's verified totals. They must agree. Returns
    {grid, review, deltas, mismatches, reconciled} or None when either side is
    absent (e.g. weekly invoice CSVs or non-template files), so callers can skip
    reconciliation without error.
    """
    grid = _grid_totals(content)
    review = _review_controls(content)
    if not grid or not review:
        return None
    metrics = ("item_count", "opening", "received", "pulled", "ending")
    deltas: dict[str, float] = {}
    mismatches: list[dict[str, float]] = []
    for m in metrics:
        if m not in review:
            continue
        g = round(grid.get(m, 0), 2)
        r = round(review[m], 2)
        deltas[m] = round(g - r, 2)
        if abs(g - r) > 0.001:
            mismatches.append(
                {"metric": m, "grid": g, "review": r, "delta": round(g - r, 2)}
            )
    parsed_financial = _parsed_financial_totals(content)
    review_financial = _review_financial_controls(content)
    financial = None
    if parsed_financial and review_financial:
        financial_metrics = (
            "opening_value",
            "received_value",
            "pulled_value",
            "ending_value",
        )
        financial_deltas = {
            m: round(parsed_financial.get(m, 0) - review_financial.get(m, 0), 2)
            for m in financial_metrics
            if m in review_financial
        }
        financial = {
            "parsed": {k: round(v, 2) for k, v in parsed_financial.items()},
            "review": {k: round(v, 2) for k, v in review_financial.items()},
            "deltas": financial_deltas,
            "reconciled": all(abs(v) <= 0.001 for v in financial_deltas.values()),
        }

    result = {
        "grid": {k: round(v, 2) for k, v in grid.items()},
        "review": {k: round(v, 2) for k, v in review.items()},
        "deltas": deltas,
        "mismatches": mismatches,
        "reconciled": len(mismatches) == 0,
        # grid is computed by applying the workbook formulas to the raw cells, so it
        # is authoritative; review controls are an advisory cross-check.
        "authoritative": "grid",
        "formulas": extract_workbook_formula_report(content),
    }
    if financial is not None:
        result["financial"] = financial
    weekly_invoice_totals = _review_weekly_invoice_totals(content)
    if weekly_invoice_totals is not None:
        result["weekly_invoice_totals"] = weekly_invoice_totals
    return result


_HEADER_ROW_TOKENS = set(_FLAT_INV_HEADER_ALIASES) | {
    "itemdescription",
    "qty",
    "quantity",
    "onhand",
    "par",
    "unit",
    "received",
    "issued",
    "pulled",
    "total",
}


def _looks_like_banner_headers(keys: list[Any]) -> bool:
    """True when a sheet's column keys are a title banner / placeholders rather
    than real headers (e.g. pandas read 'MIAMI JOB CORPS CENTER' + 'Unnamed: 1'
    ... because the real header sits a few rows down)."""
    real = [k for k in keys if k != "__sheet"]
    if not real:
        return False
    junk = 0
    recognized = 0
    for k in real:
        norm = _norm_header(k)
        low = str(k).strip().lower()
        if low.startswith(("unnamed", "col_")) or not norm:
            junk += 1
        if norm in _HEADER_ROW_TOKENS:
            recognized += 1
    # Banner if almost nothing is a recognized column AND placeholders dominate.
    return recognized < 2 and junk >= max(1, len(real) // 2)


def _reheader_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Repair rows whose header is actually a title banner.

    When a workbook carries a banner above the real header, pandas keys every
    row by that banner ('Unnamed: 1'...), burying the true column names in the
    first data row. This scans the first rows for the genuine header (a row whose
    *values* contain recognized column tokens) and re-keys every row beneath it.

    This is what lets the AI fallback succeed: rows_to_text() keys its output off
    the first row, so without this the model receives 'Unnamed: 1' columns with
    the real header as a stray data line — the exact input that was timing out.
    Re-keying also gives map_rows_to_inventory a second, deterministic chance.
    """
    if not rows or not _looks_like_banner_headers(list(rows[0].keys())):
        return rows

    header_idx: int | None = None
    for i, row in enumerate(rows[:15]):
        vals = [v for k, v in row.items() if k != "__sheet"]
        score = sum(1 for v in vals if _norm_header(v) in _HEADER_ROW_TOKENS)
        if score >= 2:
            header_idx = i
            break
    if header_idx is None:
        return rows

    header_row = rows[header_idx]
    data_keys = [k for k in header_row if k != "__sheet"]
    new_keys = [
        str(header_row[k]).strip() if header_row[k] not in (None, "") else f"col_{j}"
        for j, k in enumerate(data_keys)
    ]
    sheet = rows[0].get("__sheet")

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        vals = [row[k] for k in row if k != "__sheet"]
        if all(v is None for v in vals):
            continue
        rec: dict[str, Any] = {
            new_keys[j]: vals[j] for j in range(min(len(new_keys), len(vals)))
        }
        rec["__sheet"] = sheet
        out.append(rec)
    return out or rows


def parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def parse_tsv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    return [dict(row) for row in reader]


def parse_excel(content: bytes) -> list[dict[str, Any]]:
    monthly_inventory = _parse_mjcc_monthly_inventory(content)
    if monthly_inventory:
        return monthly_inventory

    # Flat columnar exports ("Full Inventory" / "Fact checked") carry a title
    # banner above a Category|SKU|Description|...|Ending OH|Unit Price header.
    # Detect and map these deterministically before falling back to pandas
    # (which would read the banner row as headers) or AI (which times out).
    flat_inventory = _parse_mjcc_flat_inventory(content)
    if flat_inventory:
        return flat_inventory

    try:
        import pandas as pd

        sheets = pd.read_excel(io.BytesIO(content), sheet_name=None, engine="openpyxl")
        result: list[dict[str, Any]] = []
        for sheet_name, frame in sheets.items():
            frame = frame.dropna(how="all")
            if frame.empty:
                continue
            frame.columns = [
                str(col).strip() if col is not None else f"col_{i}"
                for i, col in enumerate(frame.columns)
            ]
            sheet_rows = [
                {**row, "__sheet": sheet_name}
                for row in frame.where(pd.notnull(frame), None).to_dict(
                    orient="records"
                )
            ]
            # Repair banner-as-header sheets so both the deterministic mapper and
            # the AI fallback receive real column names instead of 'Unnamed: N'.
            result.extend(_reheader_rows(sheet_rows))
        if result:
            return result
    except ImportError:
        pass
    except Exception:
        pass

    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel parsing: pip install openpyxl"
        )
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            mapped = {
                headers[i]: row[i] if i < len(row) else None
                for i in range(len(headers))
            }
            mapped["__sheet"] = ws.title
            result.append(mapped)
    return result


_PDF_PAGE_CAP = 40  # matches invoice_parser._PDF_MAX_PAGES


def parse_pdf(content: bytes) -> str:
    pages: list[str] = []
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            total = min(len(pdf.pages), _PDF_PAGE_CAP)
            for i in range(total):
                page = pdf.pages[i]
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
                finally:
                    page.close()
                pages.append(text)
    except ImportError:
        pass

    text = "\n".join(pages).strip()
    if text:
        return text

    try:
        from pdfminer.high_level import extract_text

        return (extract_text(io.BytesIO(content), maxpages=_PDF_PAGE_CAP) or "").strip()
    except ImportError:
        raise RuntimeError("pdfplumber or pdfminer.six is required for PDF parsing.")
    except Exception:
        return ""


def rows_to_text(rows: list[dict]) -> str:
    """Convert structured rows to a readable text block for AI prompts."""
    if not rows:
        return ""
    headers = list(rows[0].keys())
    lines = ["\t".join(headers)]
    for row in rows:
        lines.append("\t".join(str(row.get(h, "")) for h in headers))
    return "\n".join(lines)


# Rotate a rendered page above this column/row profile-variance ratio. Empirically
# upright US Foods pages sit near 0.35 (col<<row) and sideways pages near 2.8-11
# (col>>row), so a 1.30 gate never fires on an already-upright page and reliably
# catches the 90-degree case. See _orient_invoice_page.
_ORIENT_ROTATE_RATIO = 1.30


def _orient_invoice_page(png_bytes: bytes) -> bytes:
    """Return the page rotated upright when its text lines run vertically.

    US Foods invoices are landscape line-item tables drawn as outlined vector
    graphics rotated 90 degrees on a portrait page (native text extraction yields
    nothing), so PyMuPDF/pdfplumber render them sideways. Both the vision model
    and the Cloud-Vision-OCR -> regex parser need upright pages, and a sideways
    render is the root cause of under-counted items and the "Page N of M" value
    landing in the invoice-number field.

    Orientation is detected with projection-profile variance in pure Pillow (no
    OCR, no extra deps): rows of text create ink banding *perpendicular* to the
    line direction, so an upright page has higher row-profile variance while a
    sideways page has higher column-profile variance. Only pages whose text is
    actually sideways are rotated, so a correctly oriented page is never touched.
    On any error the original bytes are returned unchanged.
    """
    try:
        from PIL import Image
    except Exception:
        return png_bytes
    try:
        im = Image.open(io.BytesIO(png_bytes))
        im.load()
        probe = im.convert("L")
        probe.thumbnail((400, 400))  # tiny + fast; keeps peak heap flat
        w, h = probe.size
        if w < 8 or h < 8:
            return png_bytes
        inv = probe.point(lambda p: 255 - p)  # dark ink -> high value
        row_prof = list(inv.resize((1, h)).getdata())  # mean ink per row
        col_prof = list(inv.resize((w, 1)).getdata())  # mean ink per column

        def _var(xs: list[float]) -> float:
            m = sum(xs) / len(xs)
            return sum((x - m) ** 2 for x in xs) / len(xs)

        row_var = _var(row_prof)
        col_var = _var(col_prof)
        if row_var > 0 and col_var > row_var * _ORIENT_ROTATE_RATIO:
            rotated = im.transpose(Image.ROTATE_270)
            buf = io.BytesIO()
            rotated.save(buf, format="PNG")
            return buf.getvalue()
    except Exception:
        pass
    return png_bytes


def detect_and_parse(
    filename: str, content: bytes
) -> tuple[str, list[dict] | str | dict]:
    """
    Returns (kind, data):
      'rows'           — list of dicts (CSV/Excel/TSV): deterministic column mapping.
      'text'           — plain string (PDF/txt): AI text extraction.
      'invoice_items'  — dict {'meta':..., 'items':[...]} from deterministic invoice parser.
      'invoice_images' — dict {'images':[bytes,...], 'meta':{}} for vision/OCR path.

    Content sniffed by magic bytes first; filename extension used as fallback only.
    """
    import zipfile as _zipfile

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # ── magic-byte sniffing ───────────────────────────────────────────────────
    is_pdf = content[:4] == b"%PDF"
    is_zip = content[:4] == b"PK\x03\x04"
    is_jpeg = content[:2] == b"\xff\xd8"
    is_png = content[:4] == b"\x89PNG"
    is_image = is_jpeg or is_png or f".{ext}" in invoice_parser.IMAGE_EXTENSIONS

    # ZIPs that bundle images (e.g. multi-page invoice scan saved as .pdf).
    # OOM GUARD: a 13-page JPEG bundle at native size can exceed 100-200MB
    # decompressed on a 512MB Render instance -- the same memory ceiling that
    # forced the 6-page/96dpi cap on the PDF-render fallback below. Apply the
    # same discipline here: cap pages AND re-encode/downscale each image so
    # peak heap stays bounded regardless of how the scan was captured.
    _ZIP_IMG_PAGE_CAP = 16
    _ZIP_IMG_MAX_DIM = 1600  # px, longest side
    if is_zip:
        try:
            with _zipfile.ZipFile(io.BytesIO(content)) as zf:
                all_names = [
                    n
                    for n in sorted(zf.namelist())
                    if n.lower().endswith(
                        (
                            ".jpg",
                            ".jpeg",
                            ".png",
                            ".webp",
                            ".bmp",
                            ".tif",
                            ".tiff",
                        )
                    )
                ]
                names = all_names[:_ZIP_IMG_PAGE_CAP]
                image_bytes_list: list[bytes] = []
                for name in names:
                    with zf.open(name) as img_f:
                        raw = img_f.read()
                    try:
                        from PIL import Image as _PILImage

                        _im = _PILImage.open(io.BytesIO(raw))
                        _im.load()
                        if max(_im.size) > _ZIP_IMG_MAX_DIM:
                            _scale = _ZIP_IMG_MAX_DIM / max(_im.size)
                            _im = _im.resize(
                                (
                                    max(1, int(_im.width * _scale)),
                                    max(1, int(_im.height * _scale)),
                                )
                            )
                        _buf = io.BytesIO()
                        _im.convert("RGB").save(_buf, format="JPEG", quality=82)
                        image_bytes_list.append(_buf.getvalue())
                        _im.close()
                    except Exception:
                        # If Pillow can't downscale this one, fall back to the
                        # raw bytes rather than dropping the page entirely.
                        image_bytes_list.append(raw)
                if image_bytes_list:
                    return "invoice_images", {
                        "images": image_bytes_list,
                        "meta": {
                            "filename": filename,
                            "pages_truncated": len(names) < len(all_names),
                            "pages_total": len(all_names),
                            "pages_used": len(names),
                        },
                    }
        except Exception:
            pass

    # PDFs: try deterministic invoice parser first, fall back to plain text,
    # then convert to page images for vision AI if text extraction yields nothing.
    if is_pdf or ext == "pdf":
        try:
            parsed = invoice_parser.parse_invoice_bytes_pdf(content, filename)
            if parsed["items"]:
                return "invoice_items", parsed
        except Exception:
            pass
        raw_text = ""
        try:
            raw_text = parse_pdf(content)
        except Exception:
            pass
        if raw_text.strip():
            return "text", raw_text
        # All-image PDF (scanned / no native text) — render pages as PNG for vision AI.
        # 150 DPI is the floor for the model to reliably read dense invoice line-item
        # text (product codes, packed columns, small prices). 96 DPI was tried
        # previously and reliably produced empty extractions — the model could not
        # read the text, not a timeout or API error.
        #
        # Pages are now sent to the vision model ONE AT A TIME (see
        # invoice_parser.extract_invoice_vision), not batched into one request, so
        # raising DPI does not multiply a single request's payload size — peak
        # memory is bounded by ONE page image at a time, not all pages at once.
        # We still cap total pages to guard against pathological uploads.
        _PDF_RENDER_DPI = 150
        _PDF_PAGE_CAP = 10  # ponytail: US Foods invoices ≤ 8 pages; 10 is safe ceiling
        try:
            import fitz

            page_images: list[bytes] = []
            with fitz.open(stream=content, filetype="pdf") as _doc:
                pages_total = _doc.page_count
                zoom = _PDF_RENDER_DPI / 72
                matrix = fitz.Matrix(zoom, zoom)
                for _page_index in range(min(pages_total, _PDF_PAGE_CAP)):
                    _page = _doc.load_page(_page_index)
                    _pix = _page.get_pixmap(matrix=matrix, alpha=False)
                    page_images.append(_orient_invoice_page(_pix.tobytes("png")))
            if page_images:
                return "invoice_images", {
                    "images": page_images,
                    "meta": {
                        "filename": filename,
                        "pages_total": pages_total,
                        "pages_used": len(page_images),
                        "pages_truncated": pages_total > _PDF_PAGE_CAP,
                    },
                }
        except Exception:
            pass

        try:
            import pdfplumber as _plumber

            page_images: list[bytes] = []
            with _plumber.open(io.BytesIO(content)) as _pdf:
                pages_total = len(_pdf.pages)
                for _page in _pdf.pages[:_PDF_PAGE_CAP]:
                    try:
                        _buf = io.BytesIO()
                        _page.to_image(resolution=_PDF_RENDER_DPI).save(
                            _buf, format="PNG"
                        )
                        page_images.append(_orient_invoice_page(_buf.getvalue()))
                    except Exception:
                        pass
                    finally:
                        _page.close()
            if page_images:
                return "invoice_images", {
                    "images": page_images,
                    "meta": {
                        "filename": filename,
                        "pages_total": pages_total,
                        "pages_used": len(page_images),
                        "pages_truncated": pages_total > _PDF_PAGE_CAP,
                    },
                }
        except Exception:
            pass
        return "text", ""

    # Single images: always route to vision AI (Gemini/GPT-4o/etc.) as primary.
    # Pre-running OCR here and short-circuiting on partial results caused Gemini
    # to be skipped for multi-invoice photos and thermal receipts where OCR only
    # captured some items. The OCR→regex cascade is the fallback inside _extract_ops.
    if is_image:
        try:
            content = invoice_parser._normalize_image_for_ocr(content)
        except Exception:
            pass
        return "invoice_images", {"images": [content], "meta": {"filename": filename}}

    if ext == "csv":
        return "rows", parse_csv(content)
    if ext in ("xls", "xlsx"):
        return "rows", parse_excel(content)
    if ext == "tsv":
        return "rows", parse_tsv(content)

    # Pasted / plain-text invoices (.txt) — run the deterministic invoice text
    # parser so they get the SAME structured items + reconciliation gate as PDF
    # and vision invoices. Falls back to raw text if no invoice lines are found.
    if ext == "txt":
        text = content.decode("utf-8", errors="replace")
        try:
            parsed = invoice_parser.parse_invoice_text_pages([text], filename)
            if parsed.get("items"):
                return "invoice_items", parsed
        except Exception:
            pass
        return "text", text

    # CSV heuristic for unknown text files
    try:
        rows = parse_csv(content)
        if rows and len(rows[0]) > 1:
            return "rows", rows
    except Exception:
        pass

    return "text", content.decode("utf-8", errors="replace")
